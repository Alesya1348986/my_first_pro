import openpyxl
import pretty_errors
from datetime import datetime
import time
coefficient = {20: 1.9, 15: 1.7, 10: 1.5, 5: 1.3, 1: 1.1}
slash = "|"
space = " "
# Excel
book = openpyxl.open("D:\python_pro\ebook.xlsx", read_only=True)
sheet = book.active


#Авторизация или регистрация
def reg_auto():
    r_a = input("1. Авторизация\n2. Регистрация")
    if r_a.isdigit():
        r_a = int(r_a)
        if r_a == 1:
            autorization(login, password)
        elif r_a == 2:
            registration(login, password)
        else:
            print("Неправильный ввод, попробуйте ещё раз")
            reg_auto()
    else:
        print("Для ввода необходимо ввести 1 или 2, попробуйте ещё раз")
        reg_auto()

# Регистрация
def registration(login, password):
    file1 = open("D:\python_pro\min_user.txt", "a", encoding = "utf-8")
    file1.write(login + " " + password + " " + "User")
    return "Регистрация прошла успешно"
    file1.close()
# Авторизация
def autorization(login, password):
    person  = ""
    file1 = open("D:\python_pro\min_user.txt", "r", encoding = "utf-8")
    for line in file1:
        line = line.split()
        if line[0] == login:
            if line[1] == password:
                if line[2] == "Admin":
                    person = "Admin"
                    print("Вы вошли как Admin")
                    return person
                elif line[2] == "User":
                    person = "User"
                    print("Вы вошли как User")
                    return person
            else:
                print("Неправильный пароль")
                autorization(login, password)
        else:
            print("Логина не существует")
            autorization(login, password)
    file1.close()

# Стаж + з/п
def info():
    # Расчёт стажа
    for i in range(2, sheet.max_row + 1):
        fio = sheet[i][0].value
        data = sheet[i][1].value
        salary = float(sheet[i][2].value)
        data_year = int(data[0:4])
        today = datetime.today()
        nyd = datetime(data_year, 1, 1)
        experience = int((today - nyd).days)
        if experience >= 365:
            experience = experience // 365
            return experience
        else:
            experience = 0
            return experience
#        salary_month(experience, salary)
# Расчёт зп
    def salary_month(experience, salary):
        if experience >= 20:
            salary = salary * coefficient.get(20)
            return salary
        elif experience >= 15:
            salary = salary * coefficient.get(15)
            return salary
        elif experience >=10:
            salary = salary * coefficient.get(10)
            return salary
        elif experience >= 5:
            salary = salary * coefficient.get(5)
        elif experience >= 1:
            salary = salary * coefficient.get(1)
            return salary
        else:
            return salary
# Вывод
        len_fio = 35 - len(fio)
        len_data = 14 - len(data)
        if len(str(i)) == 1:
            print(i,  fio + (" " * (len_fio)) + "|" + data + (" " * len_data) + "|" + str(salary))
        else:
            print(i, fio + (" " * (len_fio - 1)) + "|" + data + (" " * len_data) + "|" + str(salary))

# Для просмотра (User)
def read():
    for i in range(2, sheet.max_row + 1):
        fio = sheet[i][0].value
        data = sheet[i][1].value
        salary = float(sheet[i][2].value)
        len_fio = 35 - len(fio)
        len_data = 14 - len(data)
# return
        if i == 10:
            return f"{i}. {fio}{space * len_fio}{slash}{data}{space * len_data}{slash}{salary}"
        else:
            return f"{i}. {fio}{space * (len_fio - 1)}{slash}{data}{space * len_data}{slash}{salary}"

########################################################


login = input("Введите логин")
password = input("Введите пароль")
reg_auto()
# Распределение полномочий
person = autorization(login, password)
if person == "Admin":
    choice = int(input("\nВыберите действие, которое хотите просмотреть: "
                       "\n\t1. Список работников"
                       "\n\t2. Список работников с датой принятия на работу"
                       "\n\t3. Список работников и их оклад"
                       "\n\t4. Коэффициены (Зависят от стажа работы)"
                       "\n\t5. список работников и их итоговая зарплата за месяц"
                       "\n\t6. Все данные"
                       "\n\t7. Внести новые данные"
                       "\n\t0. выход\n"))
    while choice != 0:
        # 1.Спосок работников
        if choice == 1:
            print("-" * 30)
            for el in range(2, sheet.max_row + 1):
                fio = sheet[el][0].value
                print(f"{el - 1}. {fio}")
        # 2. Список работников с датой принятия на работу
        elif choice == 2:
            print("-" * 30)
            print(f"№ {space * 5}ФИО{space * 28} Дата")
            for el in range(2, sheet.max_row + 1):
                fio = sheet[el][0].value
                data = sheet[el][1].value
                len_fio = 35 - len(fio)
                if el <= 10:
                    print(f"{el - 1}. {fio}{space * len_fio}{slash} {data} ")
                else:
                    print(f"{el - 1}. {fio}{space * (len_fio - 1)}{slash} {data} ")
        # 3. Список работников и их оклад
        elif choice == 3:
            print("-" * 30)
            print(f"№ {space * 5}ФИО{space * 28} Оклад")
            for el in range(2, sheet.max_row + 1):
                fio = sheet[el][0].value
                salary = sheet[el][2].value
                len_fio = 35 - len(fio)
                if el <= 10:
                    print(f"{el - 1}. {fio}{space * len_fio}{slash} {salary}")
                else:
                    print(f"{el - 1}. {fio}{space * (len_fio - 1)}{slash} {salary}")
        #4. Коэффициены (Зависят от стажа работы)
        elif choice == 4:
            print("-" * 30)
            for key in coefficient:
               print(f"Стаж работы {key} лет и более. Коэффициент - {coefficient[key]}")
        # 5. список работников и их итоговая зарплата за месяц"
        elif choice == 5:
            print("-" * 30)
            print(f"№ {space * 5}ФИО{space * 28} Зарплата")
            for el in range(2, sheet.max_row + 1):
                fio = sheet[el][0].value
                len_fio = 35 - len(fio)
                if el <= 10:
                    print(f"{el - 1} {fio}{space * len_fio}{slash}{info()}")
                else:
                    print(f"{el - 1} {fio}{space * (len_fio - 1)}{slash}{info()}")
        # 6. Все данные"
        elif choice == 6:
            print("-" * 30)
            print(f"№ {space * 5}ФИО{space * 30} Дата:{space * 7} Оклад   Зарплата за месяц")
            for el in range(2, sheet.max_row + 1):
                fio = sheet[el][0].value
                data = sheet[el][1].value
                salary = sheet[el][2].value
                len_fio = 35 - len(fio)
                if el <= 10:
                    print(f"{el - 1}. {fio}{space * len_fio} {slash} {data} {slash} {salary} {slash} {info()}")
                else:
                    print(f"{el - 1}. {fio}{space * (len_fio - 1)} {slash} {data} {slash} {salary} {slash} {info()}")
        elif choice == 7:
            fio = input("Введите ФИО")
            data = input("Введите дату проиняти на работу в формате yyyy,mm,dd")
            salary = input("Введите оклад")
            i = [(fio, data, salary)]
            sheet.append(i)
            sheet.save("ebook.xlsx")
        choice = int(input("\nВыберите действие, которое хотите просмотреть: "
                           "\n\t1. Список работников"
                           "\n\t2. Список работников с датой принятия на работу"
                           "\n\t3. Список работников и их оклад"
                           "\n\t4. Коэффициены (Зависят от стажа работы)"
                           "\n\t5. список работников и их итоговая зарплата за месяц"
                           "\n\t6. Все данные"
                           "\n\t0. выход\n"))
# Просто чтение excel для пользователя
else:
    print(read())
book.close()












